import streamlit as st
import pandas as pd
import re
from PyPDF2 import PdfReader
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ==================================================
# Session State
# ==================================================
if "processed_df" not in st.session_state:
    st.session_state.processed_df = None

if "pdf_items_df" not in st.session_state:
    st.session_state.pdf_items_df = None

if "pdf_sa_df" not in st.session_state:
    st.session_state.pdf_sa_df = None


# ==================================================
# Helpers
# ==================================================
def to_excel_bytes(df: pd.DataFrame) -> BytesIO:
    bio = BytesIO()
    df.to_excel(bio, index=False)
    bio.seek(0)
    return bio


def normalize_headers(raw: pd.DataFrame) -> pd.DataFrame:
    header_row = None
    for i in range(min(30, len(raw))):
        if "GRADE" in raw.iloc[i].astype(str).str.upper().values:
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not locate header row containing GRADE")

    raw.columns = raw.iloc[header_row].astype(str).str.strip().str.upper()
    return raw.iloc[header_row + 1:].reset_index(drop=True)


def norm_id(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.startswith("'"):
        s = s[1:].strip()
    s = s.replace(",", "")

    # handle scientific/float-like safely
    try:
        f = float(s)
        if f.is_integer():
            s = str(int(f))
    except:
        pass

    return str(s).strip()


def norm_int_str(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.startswith("'"):
        s = s[1:].strip()
    s = s.replace(",", "")
    try:
        return str(int(float(s)))
    except:
        return ""


def series_digits_only(s: pd.Series) -> bool:
    s = s.astype(str).str.strip().replace({"": pd.NA}).dropna()
    if len(s) == 0:
        return True
    return s.str.fullmatch(r"\d+").all()


# ==================================================
# SKU Logic
# ==================================================
def map_description(grade) -> str:
    g = "" if pd.isna(grade) else str(grade).strip().upper()

    # requested behavior: blank grade -> DOG EAR
    if g == "":
        return "DOG EAR"

    if "APG" in g:
        return "TAEDA PINE APG"
    if "DOG" in g:
        return "DOG EAR"
    if re.search(r"\bIII/V\b|\bIII\b|\b3COM\b", g):
        return "TAEDA PINE #3 COMMON"

    # if it's none of the recognized ones, don't guess
    return ""


def sku_is_valid(val) -> bool:
    if pd.isna(val):
        return False
    v = str(val).strip().upper()
    return v not in ("", "NAN", "NONE")


def load_sku_lookup(sku_file) -> pd.DataFrame:
    REQUIRED = {
        "SKU": ["SKU"],
        "DESCRIPTION": ["DESCRIPTION", "DESC", "PRODUCT DESCRIPTION", "GRADE DESC"],
        "THICKNESS": ["THICKNESS", "THK", "THICK"],
        "WIDTH": ["WIDTH", "W"],
        "LENGTH": ["LENGTH", "LEN", "L"],
    }

    xls = pd.ExcelFile(sku_file)
    for sheet in xls.sheet_names:
        df = xls.parse(sheet, dtype=str)
        df.columns = df.columns.str.upper().str.strip()

        col_map = {}
        for canon, aliases in REQUIRED.items():
            for a in aliases:
                if a in df.columns:
                    col_map[a] = canon
                    break

        if set(col_map.values()) == set(REQUIRED.keys()):
            df = df.rename(columns=col_map).fillna("")
            df["DESCRIPTION"] = df["DESCRIPTION"].str.upper().str.strip()
            df["MATCH KEY"] = (
                df["DESCRIPTION"] + "|" +
                df["THICKNESS"] + "|" +
                df["WIDTH"] + "|" +
                df["LENGTH"]
            )
            return df

    raise ValueError("SKU lookup missing required columns (SKU/DESCRIPTION/THICKNESS/WIDTH/LENGTH).")


# ==================================================
# PDF Parsing (matches your receipt structure exactly)
# Receipt line format examples (from your PDFs):
#   APG 1x8x144 CHS0781 120 0.0000
#   APG 1x6x144 CHS1300 160 0.0000
#   APG 1x4x144 CHS4358 240 0.0000
# ==================================================
DIM_TOKEN = re.compile(r"^\d+\s*[xX]\s*\d+\s*[xX]\s*\d+$")
CONTAINER_TOKEN = re.compile(r"\b[A-Z]{4}\d{7}\b")


def extract_container_and_order(full_text: str, filename: str):
    # Container
    m = CONTAINER_TOKEN.search(full_text)
    container = m.group(0) if m else ""

    if not container:
        m2 = CONTAINER_TOKEN.search((filename or "").upper())
        container = m2.group(0) if m2 else ""

    # Order (from "P.O. #:" block)
    order = ""
    lines = full_text.splitlines()
    po_idx = None
    for i, line in enumerate(lines):
        if "P.O." in line.upper():
            po_idx = i
            break

    if po_idx is not None:
        for j in range(po_idx, min(po_idx + 25, len(lines))):
            for tok in re.split(r"\s+", lines[j].replace(":", " ")):
                tok = tok.strip()
                if re.fullmatch(r"\d{5,}(?:-\d+)?", tok):
                    order = tok
                    break
            if order:
                break

    # Fallback: pull from filename if present like 81804-2
    if not order:
        m3 = re.search(r"\b\d{5,}(?:-\d+)?\b", filename or "")
        if m3:
            order = m3.group(0)

    return container, order


def parse_receipt_items_from_pdfs(pdf_files) -> pd.DataFrame:
    rows = []

    for pdf in pdf_files:
        reader = PdfReader(BytesIO(pdf.getvalue()))
        pages_text = [(p.extract_text() or "") for p in reader.pages]
        full_text = "\n".join(pages_text)

        container, ordernum = extract_container_and_order(full_text, pdf.name)

        for text in pages_text:
            for line in text.splitlines():
                tokens = [t for t in line.split() if t.strip()]
                if len(tokens) < 5:
                    continue

                # find dimension token index
                dim_idx = None
                for i, tok in enumerate(tokens):
                    if DIM_TOKEN.match(tok):
                        dim_idx = i
                        break
                if dim_idx is None:
                    continue

                # Expect: [GRADE] [DIM] [LPN] [PIECES] [TOTAL_LBS...]
                if dim_idx + 2 >= len(tokens):
                    continue

                grade = tokens[dim_idx - 1] if dim_idx > 0 else ""
                dim = tokens[dim_idx]
                lpn = tokens[dim_idx + 1]
                pcs = tokens[dim_idx + 2]

                if not pcs.isdigit():
                    continue

                parts = re.split(r"[xX]", dim.replace(" ", ""))
                if len(parts) != 3:
                    continue

                try:
                    thk = int(parts[0])
                    wid = int(parts[1])
                    leng = int(parts[2])
                except:
                    continue

                pcs_int = int(pcs)
                qty_bf = pcs_int * (thk * wid * leng) / 144.0

                rows.append({
                    "PACKAGEID": str(lpn).strip(),
                    "PCS": pcs_int,
                    "QTY": qty_bf,
                    "GRADE": grade,
                    "THICKNESS": str(thk),
                    "WIDTH": str(wid),
                    "LENGTH": str(leng),
                    "CONTAINER": container,
                    "ORDERNUMBER": ordernum,
                    "PDF_FILE": pdf.name
                })

    return pd.DataFrame(rows)


def extract_lpn_pieces_map_from_pdfs(pdf_files):
    items = parse_receipt_items_from_pdfs(pdf_files)
    if items.empty:
        return set(), {}

    items["PACKAGEID"] = items["PACKAGEID"].astype(str).str.strip()
    lpns = set(items["PACKAGEID"].tolist())
    pcs_map = items.groupby("PACKAGEID")["PCS"].first().to_dict()
    pcs_map = {k: str(int(v)) for k, v in pcs_map.items()}
    return lpns, pcs_map


# ==================================================
# Full Match (Container + SKU + PDFs)
# ==================================================
def process_all(container_file, sku_file, pdf_files):
    raw_df = pd.read_excel(container_file, header=None, dtype=str)
    df = normalize_headers(raw_df).fillna("")

    required_cols = {"PACKAGEID", "PCS", "GRADE", "THICKNESS", "WIDTH", "LENGTH"}
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Container list missing required columns: {missing}")

    df["PACKAGEID"] = df["PACKAGEID"].apply(norm_id).astype(str).str.strip()
    df["PCS"] = df["PCS"].apply(norm_int_str).astype(str).str.strip()

    lpns, pcs_map = extract_lpn_pieces_map_from_pdfs(pdf_files)

    df["PDF LPN"] = df["PACKAGEID"].apply(lambda x: x if x in lpns else "")
    df["RECEIVE MATCH"] = df["PACKAGEID"].apply(lambda x: "YES" if x in lpns else "NO")

    df["PCS CHECK"] = df["PACKAGEID"].apply(lambda x: pcs_map.get(x, ""))

    def pcs_match(container_pcs, pdf_pcs):
        try:
            return "YES" if int(container_pcs) == int(pdf_pcs) else "NO"
        except:
            return "NO"

    df["PCS MATCH"] = df.apply(lambda r: pcs_match(r.get("PCS", ""), r.get("PCS CHECK", "")), axis=1)

    sku_df = load_sku_lookup(sku_file)
    df["MAPPED DESCRIPTION"] = df["GRADE"].apply(map_description)
    df["MATCH KEY"] = (
        df["MAPPED DESCRIPTION"] + "|" +
        df["THICKNESS"].astype(str) + "|" +
        df["WIDTH"].astype(str) + "|" +
        df["LENGTH"].astype(str)
    )

    df = df.merge(sku_df[["SKU", "MATCH KEY"]], how="left", on="MATCH KEY")
    df["MATCH"] = df["SKU"].apply(lambda x: "YES" if sku_is_valid(x) else "NO")
    df = df.fillna("")

    audit_cols = ["PDF LPN", "RECEIVE MATCH", "PCS CHECK", "PCS MATCH", "SKU", "MATCH"]
    existing_audit = [c for c in audit_cols if c in df.columns]
    others = [c for c in df.columns if c not in existing_audit]
    return df[others + existing_audit]


def fix_pcs_mismatch_use_container_truth(df: pd.DataFrame):
    if df is None or df.empty:
        return df, 0

    needed = {"PCS", "PCS CHECK", "PCS MATCH"}
    if not needed.issubset(set(df.columns)):
        return df, 0

    out = df.copy()
    pcs = out["PCS"].astype(str).str.strip()
    pcs_check = out["PCS CHECK"].astype(str).str.strip()
    pcs_match_col = out["PCS MATCH"].astype(str).str.upper().str.strip()

    mask = (pcs != "") & (pcs_check != "") & (pcs_match_col == "NO") & (pcs != pcs_check)

    changed = int(mask.sum())
    if changed:
        # Container is truth: update ONLY the audit fields
        out.loc[mask, "PCS CHECK"] = out.loc[mask, "PCS"]
        out.loc[mask, "PCS MATCH"] = "YES"

    return out, changed


# ==================================================
# Sales Assist Export writer (proper typing)
# ==================================================
SA_COLUMNS = [
    "SKU", "Pieces", "Quantity", "QuantityUOM", "PriceUOM", "PricePerUOM",
    "OrderNumber", "ContainerNumber", "ReloadReference", "Identifier", "ProFormaPrice"
]


def build_sales_assist_df(source_df: pd.DataFrame) -> pd.DataFrame:
    order_root = source_df.get("ORDERNUMBER", "").astype(str).str.split("-").str[0].str.strip()

    out = pd.DataFrame({
        "SKU": source_df.get("SKU", "").fillna("").astype(str),
        "Pieces": pd.to_numeric(source_df.get("PCS", 0), errors="coerce").fillna(0),
        "Quantity": pd.to_numeric(source_df.get("QTY", 0), errors="coerce").fillna(0),
        "QuantityUOM": "BF",
        "PriceUOM": "MBF",
        "PricePerUOM": 0,
        "OrderNumber": order_root.fillna("").astype(str),
        "ContainerNumber": source_df.get("CONTAINER", "").fillna("").astype(str),
        "ReloadReference": "",
        "Identifier": source_df.get("PACKAGEID", "").fillna("").astype(str),
        "ProFormaPrice": 0
    })

    out = out[SA_COLUMNS]
    return out


def sales_assist_excel_bytes(sa_df: pd.DataFrame) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(list(sa_df.columns))

    numeric_cols = {"Pieces", "Quantity", "PricePerUOM", "OrderNumber", "ProFormaPrice"}
    maybe_numeric = {"Identifier"}

    for _, row in sa_df.iterrows():
        out_row = []
        for col in sa_df.columns:
            val = row[col]

            if pd.isna(val):
                out_row.append(None)
                continue

            if col in numeric_cols:
                # OrderNumber might be digits-only or not; if not digits, keep as text
                if col == "OrderNumber":
                    s = str(val).strip()
                    if re.fullmatch(r"\d+", s):
                        out_row.append(int(s))
                    else:
                        out_row.append(s)
                else:
                    try:
                        out_row.append(float(val))
                    except:
                        out_row.append(0)

            elif col in maybe_numeric:
                s = str(val).strip()
                if re.fullmatch(r"\d+", s):
                    try:
                        out_row.append(int(s))
                    except:
                        out_row.append(s)
                else:
                    out_row.append(s)
            else:
                out_row.append(str(val))

        ws.append(out_row)

    for i, col in enumerate(sa_df.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(col) + 2))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ==================================================
# UI Styling
# ==================================================
def highlight_mismatches(row):
    if (
        row.get("RECEIVE MATCH") != "YES"
        or row.get("PCS MATCH") != "YES"
        or row.get("MATCH") != "YES"
    ):
        return ["background-color: #ffcccc"] * len(row)
    return [""] * len(row)


# ==================================================
# Streamlit UI
# ==================================================
st.set_page_config(page_title="BIFP Import Checker", layout="wide")
st.title("📦 BIFP SKU + Receive + PCS Match + Sales Assist")

container_file = st.file_uploader("Upload Container List Excel (optional)", type="xlsx")
sku_file = st.file_uploader("Upload SKU Lookup Excel", type="xlsx")
pdf_files = st.file_uploader("Upload PDF Files", type="pdf", accept_multiple_files=True)

tab1, tab2 = st.tabs(["Full Match + Audit (Container + PDFs)", "PDF + SKU Lookup → Sales Assist"])


with tab1:
    st.subheader("Full Match + Audit")

    if not (container_file and sku_file and pdf_files):
        st.info("Upload **Container List + SKU Lookup + PDFs** to run full matching.")
    else:
        if st.button("Run Full Process"):
            st.session_state.processed_df = process_all(container_file, sku_file, pdf_files)
            st.success("Full process completed")

        if st.session_state.processed_df is not None:
            dfp = st.session_state.processed_df

            try:
                st.dataframe(dfp.style.apply(highlight_mismatches, axis=1), use_container_width=True)
            except Exception:
                st.dataframe(dfp, use_container_width=True)

            c1, c2 = st.columns([1, 1])

            with c1:
                st.download_button(
                    "⬇️ Download Match Excel",
                    to_excel_bytes(dfp),
                    container_file.name.replace(".xlsx", "_SKU_RECEIVE_PCS_MATCH.xlsx"),
                )

            with c2:
                if st.button("FIX PCS Mismatch"):
                    fixed, n = fix_pcs_mismatch_use_container_truth(dfp)
                    st.session_state.processed_df = fixed
                    if n == 0:
                        st.info("No PCS mismatches found to fix (or PCS CHECK was blank).")
                    else:
                        st.success(f"Fixed {n} PCS mismatches (updated PCS CHECK + PCS MATCH only).")

            st.divider()

            st.subheader("Sales Assist Export (from Full Match)")
            sa_name = st.text_input("Sales Assist file name (no extension)", value="Sales_Assist_Full")

            if st.button("Generate Sales Assist Excel (Full Match)"):
                sa_df = build_sales_assist_df(st.session_state.processed_df)
                st.download_button(
                    "⬇️ Download Sales Assist Excel",
                    sales_assist_excel_bytes(sa_df),
                    f"{sa_name}.xlsx"
                )


with tab2:
    st.subheader("PDF + SKU Lookup → Sales Assist (no container list)")

    if not (sku_file and pdf_files):
        st.info("Upload **SKU Lookup + PDFs** to generate Sales Assist directly from PDFs.")
    else:
        if st.button("Parse PDFs + Match SKU + Build Sales Assist"):
            items_df = parse_receipt_items_from_pdfs(pdf_files)

            if items_df.empty:
                st.error(
                    "No line-items were parsed from the PDFs. "
                    "If these are scanned images, OCR would be required."
                )
            else:
                sku_df = load_sku_lookup(sku_file)

                items_df["MAPPED DESCRIPTION"] = items_df["GRADE"].apply(map_description)
                items_df["MATCH KEY"] = (
                    items_df["MAPPED DESCRIPTION"] + "|" +
                    items_df["THICKNESS"].astype(str) + "|" +
                    items_df["WIDTH"].astype(str) + "|" +
                    items_df["LENGTH"].astype(str)
                )

                items_df = items_df.merge(sku_df[["SKU", "MATCH KEY"]], how="left", on="MATCH KEY")
                items_df["MATCH"] = items_df["SKU"].apply(lambda x: "YES" if sku_is_valid(x) else "NO")
                items_df = items_df.fillna("")

                st.session_state.pdf_items_df = items_df
                st.session_state.pdf_sa_df = build_sales_assist_df(items_df)

                st.success(f"Parsed {len(items_df)} PDF line-items and built Sales Assist export.")

        if st.session_state.pdf_items_df is not None:
            st.write("Parsed PDF rows preview:")
            st.dataframe(st.session_state.pdf_items_df.head(200), use_container_width=True)

        if st.session_state.pdf_sa_df is not None:
            sa_name_pdf = st.text_input("Sales Assist file name (no extension)", value="Sales_Assist_From_PDFs")
            st.download_button(
                "⬇️ Download Sales Assist Excel (PDF-only)",
                sales_assist_excel_bytes(st.session_state.pdf_sa_df),
                f"{sa_name_pdf}.xlsx"
            )
